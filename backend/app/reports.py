"""FRS reports — build tabular report data + CSV/XLSX export, and run scheduled
reports (generate → store → email). Ported from vizor_nvr FRS (reports4.py).

Four fixed reports over a [day_from, day_to] window:
  * attendance — per person/day: First-In, Last-Out, Duration (+ face snapshot).
  * group      — per group: Headcount, Present, Attendance Compliance %.
  * mismatch   — entry/exit transit sessions: Resolved / Unresolved / Unpaired.
  * unknown    — unidentified face sightings, with detector confidence %.

Each report has a JSON shape (the ``columns``/``items`` the UI table renders) and
the same rows are exportable as CSV or XLSX via ``to_bytes``. The XLSX export
embeds the actual face snapshot thumbnail per row (fetched — and decrypted — from
object storage); CSV/JSON stay text (CSV emits a present/absent snapshot marker).
"""

from __future__ import annotations

import csv
import datetime as dt
import io

from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from edge.core.storage import _dec, get_storage

from .domain.models import Attendance, FRSEvent, Group, Person, TransitSession

REPORTS = ("attendance", "group", "mismatch", "unknown")

_XLSX_CTYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _range(day_from: str | None, day_to: str | None):
    today = dt.date.today()
    df = dt.date.fromisoformat(day_from) if day_from else today - dt.timedelta(days=7)
    dt_ = dt.date.fromisoformat(day_to) if day_to else today
    start = dt.datetime.combine(df, dt.time.min, tzinfo=dt.timezone.utc)
    end = dt.datetime.combine(dt_, dt.time.max, tzinfo=dt.timezone.utc)
    return df, dt_, start, end


def _iso(value) -> str | None:
    return value.isoformat() if value else None


def _fmt_duration(seconds: float | None) -> str:
    """Human "Xh Ym" duration; an em-dash when there's no paired out-time."""
    if not seconds or seconds < 0:
        return "—"
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    return f"{h}h {m}m"


async def build(
    db: AsyncSession,
    report: str,
    day_from: str | None,
    day_to: str | None,
    camera_ids: set | None = None,
) -> dict:
    """Build a report's tabular data.

    ``camera_ids`` optionally scopes the report to a set of cameras (C12 per-camera
    visibility). ``None`` means unrestricted (no filter). The filter is applied to
    the reports that carry a camera column (attendance, unknown); group/mismatch have
    no direct camera column and are unaffected.
    """
    df, dt_, start, end = _range(day_from, day_to)

    # ── 1. Attendance: First-In, Last-Out, Duration (per person/day) ──────────
    if report == "attendance":
        columns = ["snapshot", "day", "person_name", "first_in", "last_out", "duration"]
        stmt = (
            select(
                Attendance.day_key, Attendance.person_id, Attendance.person_name,
                Person.full_name, Attendance.check_in_at, Attendance.check_out_at,
                Attendance.check_in_snapshot, Attendance.check_out_snapshot,
            )
            .outerjoin(Person, Person.id == Attendance.person_id)
            .where(and_(Attendance.day_key >= df.isoformat(), Attendance.day_key <= dt_.isoformat()))
            .order_by(Attendance.day_key.desc(), Person.full_name)
        )
        if camera_ids is not None:
            stmt = stmt.where(Attendance.camera_id.in_(camera_ids))
        items = []
        for day, pid, aname, fname, cin, cout, cin_snap, cout_snap in (await db.execute(stmt)).all():
            last = cout or cin
            dur = (last - cin).total_seconds() if (cin and last) else None
            items.append({
                "snapshot": cin_snap or cout_snap or "",
                "day": day,
                "person_name": fname or aname or (f"Person {str(pid)[:8]}" if pid else "Unknown"),
                "first_in": _iso(cin),
                "last_out": _iso(cout) or _iso(cin),
                "duration": _fmt_duration(dur),
            })
        return {"columns": columns, "items": items, "total": len(items)}

    # ── 2. Group: Headcount, Present, Attendance Compliance % ─────────────────
    if report == "group":
        columns = ["group", "headcount", "present", "compliance_pct"]
        # Total enrolled persons per group (headcount).
        head = dict((await db.execute(
            select(Person.group_id, func.count())
            .where(Person.group_id.isnot(None))
            .group_by(Person.group_id)
        )).all())
        # Distinct persons of a group seen at least once in the window (present).
        present = dict((await db.execute(
            select(Person.group_id, func.count(func.distinct(Attendance.person_id)))
            .join(Person, Person.id == Attendance.person_id)
            .where(and_(Attendance.day_key >= df.isoformat(), Attendance.day_key <= dt_.isoformat(),
                        Person.group_id.isnot(None)))
            .group_by(Person.group_id)
        )).all())
        groups = (await db.execute(select(Group.id, Group.name))).all()
        items = []
        for gid, gname in groups:
            hc = int(head.get(gid, 0))
            pr = int(present.get(gid, 0))
            comp = round(100.0 * pr / hc, 1) if hc else 0.0
            items.append({"group": gname, "headcount": hc, "present": pr, "compliance_pct": comp})
        items.sort(key=lambda r: r["headcount"], reverse=True)
        return {"columns": columns, "items": items, "total": len(items)}

    # ── 3. Entry/Exit Mismatch: transit sessions (Resolved/Unresolved/Unpaired)
    if report == "mismatch":
        columns = ["snapshot", "person_name", "entry_time", "exit_time", "status"]
        stmt = (
            select(TransitSession, Person.full_name)
            .outerjoin(Person, Person.id == TransitSession.person_id)
            .where(and_(TransitSession.started_at >= start, TransitSession.started_at <= end))
            .order_by(TransitSession.started_at.desc())
        )
        items = []
        for sess, fname in (await db.execute(stmt)).all():
            attrs = sess.attributes or {}
            # closed = resolved (paired entry+exit); overdue/open = unresolved/unpaired.
            if sess.status == "closed":
                status = "Resolved"
            elif sess.status == "overdue":
                status = "Unresolved (overdue)"
            else:
                status = "Unpaired (no exit)"
            items.append({
                "snapshot": attrs.get("entry_snapshot") or attrs.get("exit_snapshot")
                or attrs.get("face_snapshot") or "",
                "person_name": fname or attrs.get("person_name")
                or (f"Person {str(sess.person_id)[:8]}" if sess.person_id else "Unknown"),
                "entry_time": _iso(sess.started_at),
                "exit_time": _iso(sess.ended_at) or "—",
                "status": status,
            })
        return {"columns": columns, "items": items, "total": len(items)}

    # ── 4. Unknown Attempts: time, camera, detector confidence % ──────────────
    if report == "unknown":
        # "detected_pct" is the DETECTOR confidence (a face was found) — the match
        # score is always 0 on an Unknown, so showing that reads as a confusing "0%".
        columns = ["snapshot", "time", "camera", "detected_pct"]
        stmt = (
            select(FRSEvent).where(and_(
                FRSEvent.event_type == "face_unknown",
                FRSEvent.triggered_at >= start, FRSEvent.triggered_at <= end,
            )).order_by(FRSEvent.triggered_at.desc()).limit(2000)
        )
        if camera_ids is not None:
            stmt = stmt.where(FRSEvent.camera_id.in_(camera_ids))
        evs = (await db.execute(stmt)).scalars().all()
        items = []
        for e in evs:
            attrs = e.attributes or {}
            det = attrs.get("det_confidence")
            if det is None:
                det = float(e.confidence or 0.0)
            items.append({
                "snapshot": attrs.get("face_snapshot") or e.snapshot_key or "",
                "time": _iso(e.triggered_at),
                "camera": attrs.get("camera_name") or e.camera_name
                or (str(e.camera_id)[:8] if e.camera_id else "—"),
                "detected_pct": round(float(det) * 100, 1),
            })
        return {"columns": columns, "items": items, "total": len(items)}

    raise ValueError(f"unknown report {report!r}")


# ── export ────────────────────────────────────────────────────────────────────
async def to_bytes(data: dict, fmt: str) -> tuple[bytes, str]:
    """Serialise a report dict to CSV or XLSX bytes; returns (bytes, content_type).

    CSV is plain text — it can't embed an image, so the ``snapshot`` column (when
    present) becomes a "yes"/absent marker. XLSX embeds the actual face thumbnail
    per row, fetched (and decrypted) from object storage by its stored key.
    """
    cols = data["columns"]
    items = data["items"]
    has_snap = "snapshot" in cols

    if fmt == "csv":
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for row in items:
            out = {c: row.get(c, "") for c in cols}
            if has_snap:
                out["snapshot"] = "yes" if row.get("snapshot") else ""
            w.writerow(out)
        return buf.getvalue().encode("utf-8"), "text/csv"

    # xlsx — prefetch each row's snapshot bytes (async storage read) up front, then
    # build the workbook synchronously.
    snaps: dict[str, bytes] = {}
    if has_snap:
        storage = get_storage()
        for row in items:
            key = row.get("snapshot")
            if key and key not in snaps:
                try:
                    # The S3 backend returns bytes as-stored; protected (biometric)
                    # keys are encrypted at rest, so decrypt before embedding.
                    snaps[key] = _dec(key, await storage.get(key))
                except Exception:  # noqa: BLE001 — a missing/unreadable crop just omits the image
                    snaps[key] = b""
    return _xlsx(cols, items, snaps), _XLSX_CTYPE


def _xlsx(cols: list[str], items: list[dict], snaps: dict[str, bytes]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    has_snap = "snapshot" in cols
    ws.append([c.replace("_", " ").title() for c in cols])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")

    THUMB = 56  # px — embedded face thumbnail size
    for ri, row in enumerate(items, start=2):
        ws.append([("" if c == "snapshot" else row.get(c, "")) for c in cols])
        if not has_snap:
            continue
        raw = snaps.get(row.get("snapshot") or "")
        if not raw:
            continue
        try:
            img = XLImage(io.BytesIO(raw))
            img.width = img.height = THUMB
            col_letter = get_column_letter(cols.index("snapshot") + 1)
            ws.row_dimensions[ri].height = THUMB * 0.78  # pt
            ws.add_image(img, f"{col_letter}{ri}")
        except Exception:  # noqa: BLE001 — bad bytes just leave the cell empty
            pass

    if has_snap:
        ws.column_dimensions[get_column_letter(cols.index("snapshot") + 1)].width = 10

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def next_run(frequency: str, at_time: str, after: dt.datetime | None = None) -> dt.datetime:
    """Compute the next fire time for a schedule (UTC)."""
    after = after or dt.datetime.now(dt.timezone.utc)
    try:
        hh, mm = (int(x) for x in at_time.split(":"))
    except ValueError:
        hh, mm = 8, 0
    nxt = after.replace(hour=hh, minute=mm, second=0, microsecond=0)
    if nxt <= after:
        nxt += dt.timedelta(days=1)
    step = {"daily": 1, "weekly": 7, "monthly": 30}.get(frequency, 1)
    # For weekly/monthly, roll forward to at least `step` days out from `after`.
    while (nxt - after).days < (step - 1):
        nxt += dt.timedelta(days=1)
    return nxt


async def run_schedule(db: AsyncSession, schedule) -> "object":
    """Generate a schedule's report, store the file, email recipients, record a run."""
    from edge.core.storage import get_storage

    from .domain.models import ReportRun

    df = (dt.date.today() - dt.timedelta(days=schedule.range_days)).isoformat()
    dt_ = dt.date.today().isoformat()
    data = await build(db, schedule.report, df, dt_)
    blob, ctype = await to_bytes(data, schedule.fmt)
    stamp = dt.datetime.now(dt.timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"{schedule.report}-{stamp}.{schedule.fmt}"
    key = f"frs/reports/{filename}"
    await get_storage().put(key, blob, ctype)

    emailed, ok = None, False
    recipients = [r.strip() for r in (schedule.recipients or "").split(",") if r.strip()]
    if recipients:
        try:
            from edge.core.config import get_settings
            from edge.messaging.email import send_email

            link = f"{get_settings().frontend_url.rstrip('/')}/reports"
            html = (
                f"<p>Your <strong>{schedule.report}</strong> report for {df} → {dt_} is ready "
                f"({len(data['items'])} rows).</p><p>Download it from the Reports page: "
                f'<a href="{link}">{link}</a></p>'
            )
            await send_email(db, recipients, f"FRS {schedule.report} report", html)
            emailed, ok = ", ".join(recipients), True
        except Exception:  # noqa: BLE001 — email is best-effort
            emailed, ok = ", ".join(recipients), False

    run = ReportRun(schedule_id=schedule.id, report=schedule.report, fmt=schedule.fmt,
                    filename=filename, path=key, rows=len(data["items"]), emailed_to=emailed, email_ok=ok)
    db.add(run)
    schedule.last_run_at = dt.datetime.now(dt.timezone.utc)
    schedule.next_run_at = next_run(schedule.frequency, schedule.at_time)
    await db.commit()
    await db.refresh(run)
    return run
