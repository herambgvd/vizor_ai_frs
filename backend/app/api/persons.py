"""Person gallery CRUD + bulk import + government-ID documents.

Ported from vizor_nvr FRS `routers/persons.py`, re-implemented on the platform-edge
stack. Face photos + their embeddings are added by the Photos port; deleting a
person here performs a GDPR erasure of everything currently held (profile + ID
document), extended to purge photos/events/vectors as those features land.
"""

from __future__ import annotations

import datetime as dt
import io
import os
import uuid

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from edge.auth.deps import require_permission
from edge.core.audit import record as audit_record
from edge.core.errors import ConflictError, NotFoundError, ValidationError
from edge.core.pagination import Page, PageParams, page_params, paginate
from edge.core.storage import get_storage
from edge.db.base import get_db

from ..domain.models import Group, Person
from ..domain.permissions import FrsPerm
from .schemas import PersonCreate, PersonOut, PersonUpdate

router = APIRouter(prefix="/frs/persons", tags=["frs-persons"])

# Bulk-import columns (header row). full_name is the only required column.
_IMPORT_COLUMNS = [
    "full_name", "external_id", "group", "category", "priority",
    "department", "designation", "contact_number", "date_of_joining",
    "id_type", "id_number", "validity_start", "validity_end", "auto_remove",
]
_MAX_IMPORT_ROWS = 2000


def _person_out(p: Person) -> PersonOut:
    out = PersonOut.model_validate(p)
    out.has_id_document = p.id_file_key is not None
    return out


# --- list + create -----------------------------------------------------------
@router.get("", response_model=Page[PersonOut])
async def list_persons(
    search: str | None = Query(default=None, description="Match full_name or external_id"),
    group_id: uuid.UUID | None = Query(default=None),
    category: str | None = Query(default=None),
    params: PageParams = Depends(page_params),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission(FrsPerm.PERSON_READ)),
) -> Page[PersonOut]:
    stmt = select(Person).order_by(Person.full_name)
    if search:
        like = f"%{search}%"
        stmt = stmt.where(or_(Person.full_name.ilike(like), Person.external_id.ilike(like)))
    if group_id is not None:
        stmt = stmt.where(Person.group_id == group_id)
    if category:
        stmt = stmt.where(Person.category == category)
    page = await paginate(db, stmt, params)
    page.items = [_person_out(p) for p in page.items]
    return page


async def _ensure_unique_external_id(db: AsyncSession, external_id: str | None, exclude_id=None) -> None:
    if not external_id:
        return
    stmt = select(Person).where(Person.external_id == external_id)
    if exclude_id is not None:
        stmt = stmt.where(Person.id != exclude_id)
    if (await db.execute(stmt)).scalar_one_or_none():
        raise ConflictError(f"a person with external_id {external_id!r} already exists")


@router.post("", response_model=PersonOut, status_code=201)
async def create_person(
    data: PersonCreate,
    db: AsyncSession = Depends(get_db),
    actor=Depends(require_permission(FrsPerm.PERSON_MANAGE)),
) -> PersonOut:
    await _ensure_unique_external_id(db, data.external_id)
    if data.group_id is not None and await db.get(Group, data.group_id) is None:
        raise ValidationError("group not found")
    p = Person(**data.model_dump())
    db.add(p)
    await db.commit()
    await db.refresh(p)
    await audit_record(
        db, actor=actor, action="frs.person.create", target_type="frs_person",
        target_id=str(p.id), meta={"full_name": p.full_name},
    )
    return _person_out(p)


# --- bulk import (declared before /{person_id} so the static paths win) -------
@router.get("/import-template")
async def import_template(
    _=Depends(require_permission(FrsPerm.PERSON_MANAGE)),
) -> StreamingResponse:
    """Download an XLSX template with the bulk-import column headers."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "persons"
    ws.append(_IMPORT_COLUMNS)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=persons-import-template.xlsx"},
    )


def _coerce_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return dt.date.fromisoformat(str(value).strip()[:10])


def _truthy(value) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _parse_import(raw: bytes, filename: str) -> list[dict]:
    """Read the uploaded XLSX/CSV into a list of column→value dicts (header-mapped)."""
    name = (filename or "").lower()
    rows: list[list] = []
    if name.endswith(".csv"):
        import csv

        text = raw.decode("utf-8-sig", errors="replace")
        rows = [r for r in csv.reader(io.StringIO(text))]
    else:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    header = [str(c or "").strip().lower() for c in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(c in (None, "") for c in r):
            continue
        out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    return out


@router.post("/import")
async def import_persons(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    actor=Depends(require_permission(FrsPerm.PERSON_MANAGE)),
) -> dict:
    """Bulk create/update persons from an XLSX/CSV. Upserts by external_id; resolves
    group by name. Photos are enrolled separately (Photos port)."""
    records = _parse_import(await file.read(), file.filename or "")
    if len(records) > _MAX_IMPORT_ROWS:
        raise ValidationError(f"import is limited to {_MAX_IMPORT_ROWS} rows")

    groups = (await db.execute(select(Group))).scalars().all()
    group_by_name = {g.name.strip().lower(): g for g in groups}

    created, updated, skipped, errors = 0, 0, 0, []
    for i, row in enumerate(records, start=2):  # row 1 is the header
        full_name = str(row.get("full_name") or "").strip()
        if not full_name:
            skipped += 1
            continue
        ext = str(row.get("external_id") or "").strip() or None
        gname = str(row.get("group") or "").strip().lower()
        group = group_by_name.get(gname)
        try:
            fields = dict(
                full_name=full_name,
                external_id=ext,
                group_id=group.id if group else None,
                category=str(row.get("category") or "").strip() or None,
                priority=int(row.get("priority") or 0),
                department=str(row.get("department") or "").strip() or None,
                designation=str(row.get("designation") or "").strip() or None,
                contact_number=str(row.get("contact_number") or "").strip() or None,
                date_of_joining=_coerce_date(row.get("date_of_joining")),
                id_type=str(row.get("id_type") or "").strip() or None,
                id_number=str(row.get("id_number") or "").strip() or None,
                validity_start=_coerce_date(row.get("validity_start")),
                validity_end=_coerce_date(row.get("validity_end")),
                auto_remove=_truthy(row.get("auto_remove", "")),
            )
            existing = None
            if ext:
                existing = (
                    await db.execute(select(Person).where(Person.external_id == ext))
                ).scalar_one_or_none()
            if existing:
                for key, value in fields.items():
                    setattr(existing, key, value)
                updated += 1
            else:
                db.add(Person(**fields))
                created += 1
        except Exception as exc:  # noqa: BLE001 — collect per-row errors, keep going
            skipped += 1
            errors.append({"row": i, "full_name": full_name, "error": str(exc)})
    await db.commit()
    await audit_record(
        db, actor=actor, action="frs.person.import", target_type="frs_person",
        target_id="bulk", meta={"created": created, "updated": updated, "skipped": skipped},
    )
    return {"total": len(records), "created": created, "updated": updated, "skipped": skipped, "errors": errors[:25]}


# --- single person -----------------------------------------------------------
@router.get("/{person_id}", response_model=PersonOut)
async def get_person(
    person_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission(FrsPerm.PERSON_READ)),
) -> PersonOut:
    p = await db.get(Person, person_id)
    if p is None:
        raise NotFoundError("person not found")
    return _person_out(p)


@router.put("/{person_id}", response_model=PersonOut)
async def update_person(
    person_id: uuid.UUID,
    data: PersonUpdate,
    db: AsyncSession = Depends(get_db),
    actor=Depends(require_permission(FrsPerm.PERSON_MANAGE)),
) -> PersonOut:
    p = await db.get(Person, person_id)
    if p is None:
        raise NotFoundError("person not found")
    patch = data.model_dump(exclude_unset=True)
    if "external_id" in patch:
        await _ensure_unique_external_id(db, patch["external_id"], exclude_id=person_id)
    if patch.get("group_id") is not None and await db.get(Group, patch["group_id"]) is None:
        raise ValidationError("group not found")
    for key, value in patch.items():
        setattr(p, key, value)
    await db.commit()
    await db.refresh(p)
    await audit_record(
        db, actor=actor, action="frs.person.update", target_type="frs_person",
        target_id=str(person_id), meta={k: str(v) for k, v in patch.items()},
    )
    return _person_out(p)


@router.delete("/{person_id}", status_code=204)
async def delete_person(
    person_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    actor=Depends(require_permission(FrsPerm.PERSON_MANAGE)),
) -> None:
    """Right-to-erasure (GDPR/DPDP). Erases EVERY trace of a person's biometrics —
    mirrors vizor_nvr ``purge_person_biometrics``:

      * gallery vectors (enrolled-face embeddings) — ``gallery.delete_by_person``
      * live-sighting / snapshot vectors — ``gallery.delete_snapshot`` per event
      * their ``FRSEvent`` rows + the stored face-crop blobs
      * their ``Attendance`` rows
      * their ``Photo`` rows + the stored image/thumbnail blobs
      * the person's ID-document blob + profile thumbnail blob
      * the person row itself

    Idempotent (missing blobs / already-gone vectors are ignored). References from
    TransitSession / feedback are left as SET NULL by their FK, as intended.
    """
    import contextlib

    from fastapi.concurrency import run_in_threadpool

    from .. import gallery
    from ..domain.models import Attendance, FRSEvent, Photo

    p = await db.get(Person, person_id)
    if p is None:
        raise NotFoundError("person not found")
    name, id_key, thumb_key = p.full_name, p.id_file_key, p.thumbnail_key
    pid = str(person_id)

    # 1. Load the person's events first — we need their ids (for snapshot vectors)
    #    and their snapshot blob keys before the rows are deleted.
    events = (
        await db.execute(select(FRSEvent).where(FRSEvent.person_id == person_id))
    ).scalars().all()
    event_snapshot_keys: list[str] = []
    for ev in events:
        attrs = ev.attributes or {}
        for k in (ev.snapshot_key, attrs.get("face_snapshot")):
            if k:
                event_snapshot_keys.append(k)

    # 2. Load photo blob keys before deleting the rows.
    photos = (
        await db.execute(select(Photo).where(Photo.person_id == person_id))
    ).scalars().all()
    photo_keys: list[str] = []
    for ph in photos:
        for k in (ph.storage_key, ph.thumbnail_key):
            if k:
                photo_keys.append(k)

    # 3. Purge vectors (gallery + forensic snapshots). Qdrant calls are sync.
    await run_in_threadpool(gallery.delete_by_person, pid)
    for ev in events:
        await run_in_threadpool(gallery.delete_snapshot, str(ev.id))

    # 4. Delete the DB rows: events + attendance + photos explicitly (GDPR — not the
    #    FK's SET NULL/CASCADE default), then the person.
    for ev in events:
        await db.delete(ev)
    for a in (
        await db.execute(select(Attendance).where(Attendance.person_id == person_id))
    ).scalars().all():
        await db.delete(a)
    for ph in photos:
        await db.delete(ph)
    await db.delete(p)
    await db.commit()

    # 5. Delete the stored blobs (best-effort, idempotent).
    for key in [*photo_keys, *event_snapshot_keys, id_key, thumb_key]:
        if not key:
            continue
        with contextlib.suppress(Exception):
            await get_storage().delete(key)

    await audit_record(
        db, actor=actor, action="frs.person.delete", target_type="frs_person",
        target_id=pid, meta={"full_name": name},
    )


# --- government ID document ---------------------------------------------------
@router.post("/{person_id}/id-document", response_model=PersonOut)
async def upload_id_document(
    person_id: uuid.UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    actor=Depends(require_permission(FrsPerm.PERSON_MANAGE)),
) -> PersonOut:
    """Upload a government ID (image/PDF). Stored under the encrypted-at-rest frs/ prefix."""
    p = await db.get(Person, person_id)
    if p is None:
        raise NotFoundError("person not found")
    data = await file.read()
    ext = os.path.splitext(file.filename or "")[1]
    key = f"frs/persons/{p.id}/id_{uuid.uuid4().hex}{ext}"
    await get_storage().put(key, data, file.content_type)
    old = p.id_file_key
    p.id_file_key = key
    await db.commit()
    await db.refresh(p)
    if old and old != key:
        try:
            await get_storage().delete(old)
        except Exception:  # pragma: no cover
            pass
    await audit_record(
        db, actor=actor, action="frs.person.id_upload", target_type="frs_person",
        target_id=str(person_id),
    )
    return _person_out(p)


@router.get("/{person_id}/id-document")
async def get_id_document(
    person_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission(FrsPerm.PERSON_READ)),
) -> dict:
    """Return a fetchable URL for the person's stored ID document."""
    p = await db.get(Person, person_id)
    if p is None or not p.id_file_key:
        raise NotFoundError("no ID document for this person")
    return {"url": await get_storage().url(p.id_file_key)}


@router.delete("/{person_id}/id-document", response_model=PersonOut)
async def delete_id_document(
    person_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    actor=Depends(require_permission(FrsPerm.PERSON_MANAGE)),
) -> PersonOut:
    p = await db.get(Person, person_id)
    if p is None:
        raise NotFoundError("person not found")
    key = p.id_file_key
    p.id_file_key = None
    await db.commit()
    await db.refresh(p)
    if key:
        try:
            await get_storage().delete(key)
        except Exception:  # pragma: no cover
            pass
    await audit_record(
        db, actor=actor, action="frs.person.id_delete", target_type="frs_person",
        target_id=str(person_id),
    )
    return _person_out(p)
